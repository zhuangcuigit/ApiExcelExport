#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
API 数据导出 Excel 工具
根据 config.yaml 中的接口地址、参数和列配置，请求接口并导出为 Excel。
"""

import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

import requests
import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def load_config(config_path: str = "config.yaml") -> dict:
    """加载配置文件."""
    path = Path(config_path)
    if not path.exists():
        print(f"错误：配置文件不存在 {path}")
        sys.exit(1)
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def get_nested(data, path: str):
    """按点号路径从对象中取值，如 'data.list' -> data['list']，'member.salesmanName' -> member['salesmanName']."""
    if not path or not path.strip():
        return data
    keys = path.strip().split(".")
    for k in keys:
        if data is None:
            return None
        if isinstance(data, dict) and k in data:
            data = data[k]
        else:
            return None
    return data


def fetch_data(api_config: dict, session: Optional[requests.Session] = None) -> dict:
    """请求接口并返回 JSON。若传入 session 则用 session 发请求，并把 session 的请求头（如 Cookie、x-access-token）合并进本次请求头."""
    url = api_config.get("url")
    if not url:
        raise ValueError("config.api.url 不能为空")
    method = (api_config.get("method") or "GET").upper()
    headers = dict(api_config.get("headers") or {})
    if session is None:
        cookie = api_config.get("cookie") or os.environ.get("COOKIE")
        if cookie and isinstance(cookie, str) and cookie.strip():
            headers["Cookie"] = cookie.strip()
    else:
        # 版本2：把登录后的 session 请求头（Cookie、x-access-token 等）合并进数据接口请求头
        for k, v in session.headers.items():
            if v is not None and str(v).strip():
                headers[k] = v
    params = api_config.get("params") or {}
    body = api_config.get("body")
    print("headers", headers)

    req = requests if session is None else session
    if method == "GET":
        resp = req.get(url, params=params, headers=headers, timeout=30)
    elif method == "POST":
        payload = body if body is not None else params
        resp = req.post(url, json=payload, headers=headers, timeout=30)
    else:
        raise ValueError(f"不支持的 method: {method}")

    if not resp.ok:
        print("响应内容：", resp.text[:500] if resp.text else "(无内容)")
        if resp.status_code == 400:
            print("提示：400 多为请求方式或参数不符。若该接口要求 POST 或参数名不同，请修改 config：api.method 改为 POST，或调整 api.params / api.body")
        resp.raise_for_status()

    return resp.json()


def extract_list(full_data: dict, data_path: str) -> list:
    """从完整响应中按 data_path 取出列表；若为数组则直接返回."""
    if not data_path or not data_path.strip():
        if isinstance(full_data, list):
            return full_data
        return [full_data]
    lst = get_nested(full_data, data_path)
    if lst is None:
        return []
    if not isinstance(lst, list):
        return [lst]
    return lst


def get_cell_value(row: dict, field: str):
    """从行对象中按路径取值，支持任意层级如 user.name、member.salesmanName."""
    if not field or not field.strip():
        return ""
    v = get_nested(row, field)
    return v if v is not None else ""


def _normalize_iso_fraction(s: str) -> str:
    """将 ISO 字符串中的小数秒截断为最多 6 位（strftime %f 只支持微秒）."""
    m = re.search(r"(\.\d+)", s)
    if m:
        frac = m.group(1)
        if len(frac) > 7:
            s = s[: m.start()] + frac[:7] + s[m.end() :]
    return s


def _parse_datetime(value):
    """将接口返回值解析为 datetime；支持 ISO 字符串、毫秒/秒时间戳、纯数字字符串."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)):
        try:
            ts = float(value)
            if ts > 1e12:
                ts = ts / 1000.0
            return datetime.fromtimestamp(ts)
        except (OSError, ValueError, OverflowError):
            return None
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        # 纯数字字符串视为时间戳（毫秒或秒）
        if value.isdigit() or (value.startswith("-") and value[1:].isdigit()):
            try:
                ts = int(value)
                if abs(ts) > 1e12:
                    ts = ts / 1000.0
                return datetime.fromtimestamp(ts)
            except (OSError, ValueError, OverflowError):
                pass
        # 先尝试 fromisoformat（Python 3.7+），兼容常见 ISO 格式
        normalized = value.replace("Z", "+00:00").replace("z", "+00:00")
        normalized = _normalize_iso_fraction(normalized)
        try:
            dt = datetime.fromisoformat(normalized)
            return dt.replace(tzinfo=None) if dt.tzinfo else dt
        except (ValueError, TypeError):
            pass
        # 带小数秒的格式：先截断为 6 位再试
        with_frac = _normalize_iso_fraction(value.replace("Z", "").replace("z", ""))
        for fmt in ("%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%d %H:%M:%S.%f"):
            try:
                return datetime.strptime(with_frac[:26], fmt)
            except (ValueError, TypeError):
                continue
        # 去掉时区与小数秒后再用 strptime
        stripped = re.sub(r"[+-]\d{2}:?\d{2}$", "", value)
        stripped = re.sub(r"Z$", "", stripped, flags=re.I)
        stripped = re.sub(r"\.\d+Z?$", "", stripped, flags=re.I)
        stripped = stripped.strip()
        for fmt, s in (
            ("%Y-%m-%dT%H:%M:%S", stripped),
            ("%Y-%m-%d %H:%M:%S", stripped),
            ("%Y-%m-%d", stripped[:10]),
            ("%Y/%m/%d %H:%M:%S", stripped),
            ("%Y/%m/%d", stripped[:10]),
            ("%Y%m%d%H%M%S", stripped),
            ("%Y%m%d", stripped[:8]),
        ):
            if not s:
                continue
            try:
                return datetime.strptime(s, fmt)
            except (ValueError, TypeError):
                continue
    return None


def format_cell_value(value, format_str: Optional[str]):
    """按格式串格式化单元格值；若为时间则按 format_str（strftime）格式化，否则原样返回."""
    if format_str is None or not str(format_str).strip():
        return value if value != "" else ""
    if value is None or value == "":
        return ""
    fmt = str(format_str).strip()
    dt = _parse_datetime(value)
    if dt is not None:
        try:
            return dt.strftime(fmt)
        except (ValueError, TypeError):
            return str(value)
    return value


def normalize_columns(columns: dict) -> list:
    """将 columns 配置规范为 [(field, header, format_str), ...]，支持简写与对象写法."""
    result = []
    for field, spec in columns.items():
        if isinstance(spec, str):
            result.append((field, spec, None))
        elif isinstance(spec, dict):
            header = spec.get("header") or spec.get("name") or field
            result.append((field, header, spec.get("format")))
        else:
            result.append((field, str(spec), None))
    return result


def _is_empty(value) -> bool:
    """判断是否为空：None、空字符串、仅空白."""
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def apply_filters(rows: list, filters: list) -> list:
    """按配置的过滤规则过滤行；不配置或为空则返回原列表."""
    if not filters or not isinstance(filters, list):
        return rows
    result = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        keep = True
        for rule_cfg in filters:
            if not isinstance(rule_cfg, dict):
                continue
            column = rule_cfg.get("column")
            rule = rule_cfg.get("rule")
            if not column or not rule:
                continue
            val = get_cell_value(row, column)
            if rule == "not_empty":
                if _is_empty(val):
                    keep = False
                    break
            elif rule == "equals":
                if val != rule_cfg.get("value"):
                    keep = False
                    break
            elif rule == "in":
                if val not in rule_cfg.get("values", []):
                    keep = False
                    break
        if keep:
            result.append(row)
    return result


def row_to_values(row: dict, columns_normalized: list) -> list:
    """根据列配置从一行对象中按路径取值，并应用时间格式."""
    return [
        format_cell_value(get_cell_value(row, field), fmt)
        for field, _header, fmt in columns_normalized
    ]


def export_excel(rows: list, columns_normalized: list, output_path: str) -> None:
    """将数据按列配置写入 Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "数据"

    headers = [header for _f, header, _fmt in columns_normalized]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        if not isinstance(row, dict):
            continue
        ws.append(row_to_values(row, columns_normalized))

    # 自动列宽（简单按内容长度）
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    wb.save(output_path)
    print(f"已导出: {output_path}")


def main():
    config = load_config()
    api_cfg = config.get("api") or {}
    data_path = config.get("data_path") or ""
    columns = config.get("columns") or {}
    out_cfg = config.get("output") or {}
    out_dir = out_cfg.get("dir") or "./output"
    out_name = out_cfg.get("filename") or "export_data"

    if not columns:
        print("错误：请在 config.yaml 中配置 columns")
        sys.exit(1)

    columns_normalized = normalize_columns(columns)
    filters = config.get("filters") or []

    print("正在请求接口...")
    full_data = fetch_data(api_cfg)
    # print("响应内容：", full_data)
    # print("data_path：", data_path)
    print("返回结果提示信息：", full_data.get("message"))
    rows = extract_list(full_data, data_path)
    print(f"获取到 {len(rows)} 条数据")

    if filters:
        before = len(rows)
        rows = apply_filters(rows, filters)
        print(f"过滤后保留 {len(rows)} 条（过滤掉 {before - len(rows)} 条）")

    if not rows:
        print("没有数据可导出")
        sys.exit(0)

    Path(out_dir).mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(out_dir, f"{out_name}_{ts}.xlsx")
    export_excel(rows, columns_normalized, out_path)


if __name__ == "__main__":
    main()
